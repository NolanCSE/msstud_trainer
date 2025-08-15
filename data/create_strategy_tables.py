import re
import pandas as pd
from core.simulation import run_simulation
from io import StringIO
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

strategies = ['basic', 'ap3', 'ap5']
antes = [5, 10, 15, 20]
bankrolls = [10000]
rounds_per_hour_list = [20, 30, 40, 50]  # Adjust as needed
rounds = 5000000  # bump as you like

# EXACT columns we’ll output (N0 removed)
fieldnames = [
    'Strategy','Ante','Bankroll','Rounds/Hour',
    'EV/hand','EV/hr','Std Dev','Win %','Loss %','Push %','RoR',
    'Avg Total Bet','μ (risk units)','σ (risk units)'
]
results = []

_float = r'[-+]?(?:\d+(?:\.\d*)?|\.\d+)'
def parse_metrics(text: str) -> dict:
    # Multiline, case-insensitive
    flags = re.IGNORECASE | re.MULTILINE

    def grab(pat):
        m = re.search(pat, text, flags)
        return m.group(1) if m else None

    # Pull values using the *printed* labels from run_simulation
    data = {
        'EV/hand':        grab(r'^\s*EV per hand:\s*\$(' + _float + r')'),
        'EV/hr':          grab(r'^\s*EV/hr:\s*\$(' + _float + r')'),
        'Std Dev':        grab(r'^\s*Standard Deviation:\s*\$(' + _float + r')'),
        'Win %':          grab(r'^\s*Win Rate:\s*(' + _float + r')%'),
        'Loss %':         grab(r'^\s*Loss Rate:\s*(' + _float + r')%'),
        'Push %':         grab(r'^\s*Push Rate:\s*(' + _float + r')%'),
        'RoR':            grab(r'^\s*Risk of Ruin .*:\s*(' + _float + r')%'),
        'Avg Total Bet':  grab(r'^\s*Avg total bet T̄:\s*\$(' + _float + r')'),
        # μ and σ are printed on the same line; match both
        'μ (risk units)': grab(r'^\s*μ \(risk units\):\s*(' + _float + r')'),
        'σ (risk units)': grab(r'σ \(risk units\):\s*(' + _float + r')'),
    }
    return data

def main():
    for strategy in strategies:
        for ante in antes:
            for bankroll in bankrolls:
                for rph in rounds_per_hour_list:
                    print(f"Simulating {strategy} | Ante: {ante} | Bankroll: {bankroll} | RPH: {rph}")
                    buf = StringIO()

                    # Capture run_simulation stdout
                    import sys
                    old = sys.stdout
                    sys.stdout = buf
                    try:
                        run_simulation(strategy, rounds, ante, bankroll, verbose=False, rounds_per_hour=rph)
                    finally:
                        sys.stdout = old

                    out = buf.getvalue()
                    metrics = parse_metrics(out)

                    # Optional: sanity-check first few runs
                    # print(out); print(metrics)

                    results.append({
                        'Strategy': strategy,
                        'Ante': ante,
                        'Bankroll': bankroll,
                        'Rounds/Hour': rph,
                        'EV/hand': metrics.get('EV/hand') or '',
                        'EV/hr': metrics.get('EV/hr') or '',
                        'Std Dev': metrics.get('Std Dev') or '',
                        'Win %': metrics.get('Win %') or '',
                        'Loss %': metrics.get('Loss %') or '',
                        'Push %': metrics.get('Push %') or '',
                        'RoR': metrics.get('RoR') or '',
                        'Avg Total Bet': metrics.get('Avg Total Bet') or '',
                        'μ (risk units)': metrics.get('μ (risk units)') or '',
                        'σ (risk units)': metrics.get('σ (risk units)') or '',
                    })

    # Build DataFrame in the exact order we expect
    df = pd.DataFrame(results, columns=fieldnames)

    # Write to Excel
    out_xlsx = "10k_strategy_list_mississippi_stud_strategy_analysis.xlsx"
    with pd.ExcelWriter(out_xlsx, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Simulation Results')
        ws = writer.sheets['Simulation Results']
        for i, title in enumerate(df.columns, 1):
            col = get_column_letter(i)
            ws.column_dimensions[col].width = max(14, len(title) + 2)
            cell = ws[f"{col}1"]
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
    print(f"\n✅ Done. Results saved to '{out_xlsx}'")

if __name__ == '__main__':
    import multiprocessing
    multiprocessing.freeze_support()
    main()
